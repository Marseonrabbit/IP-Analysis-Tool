import sys
import os
import threading
import requests
import pandas as pd
import webbrowser
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTabWidget, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QPushButton, QTextEdit, 
                             QFileDialog, QProgressBar, QMessageBox)
from PyQt5.QtGui import QFont
from PyQt5.QtCore import Qt
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

downloads_folder = os.path.join(os.path.expanduser("~"), "Downloads")
selected_file_path = ""
api_key_file = "api_key.txt"

def get_country_name(country_code):
    if not country_code:
        return "Unknown Country"
    try:
        response = requests.get(f"https://restcountries.com/v3.1/alpha/{country_code}")
        if response.status_code == 200:
            country_data = response.json()
            return country_data[0]["name"]["common"]
        else:
            return "Unknown Country"
    except:
        return "Unknown Country"

def get_ip_info(api_key, ip):
    url = f"https://www.virustotal.com/api/v3/ip_addresses/{ip}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)

    if response.status_code == 401:
        return None, None, None

    if response.status_code == 200:
        data = response.json()
        isp = data.get("data", {}).get("attributes", {}).get("as_owner", "Unknown ISP")
        country_code = data.get("data", {}).get("attributes", {}).get("country", "")
        country = get_country_name(country_code)
        malicious_count = data.get("data", {}).get("attributes", {}).get("last_analysis_stats", {}).get("malicious", 0)
        reputation = classify_reputation(malicious_count)
        return isp, country, reputation
    else:
        return "Error", "Error", "Error"

def classify_reputation(score):
    if score == 0:
        return "Safe"
    elif 1 <= score <= 10:
        return "Neutral"
    elif score > 10:
        return "Poor"
    return "Unknown"

def get_hash_reputation(api_key, file_hash):
    url = f"https://www.virustotal.com/api/v3/files/{file_hash}"
    headers = {"x-apikey": api_key}
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        data = response.json()
        attributes = data.get("data", {}).get("attributes", {})

        # Get the required data
        last_analysis_stats = attributes.get("last_analysis_stats", {})
        malicious_count = last_analysis_stats.get("malicious", 0)
        total_engines = sum(last_analysis_stats.values())
        reputation = classify_hash_reputation(malicious_count, total_engines)

        # Get the community score
        community_score = attributes.get("reputation", 0)

        # Get file signature info
        signature_info = attributes.get("signature_info", {})
        file_signer = signature_info.get("subject", "Unknown")

        # Other details
        file_type = attributes.get("type_description", "Unknown")
        file_size = attributes.get("size", "Unknown")
        file_names = attributes.get("names", [])

        return reputation, malicious_count, total_engines, community_score, file_signer, file_type, file_size, file_names
    elif response.status_code == 404:
        return "Not Found", 0, 0, 0, "Unknown", "Unknown", "Unknown", []
    elif response.status_code == 401:
        return "Invalid API Key", 0, 0, 0, "Unknown", "Unknown", "Unknown", []
    else:
        return "Error", 0, 0, 0, "Unknown", "Unknown", "Unknown", []

def classify_hash_reputation(malicious_count, total_engines):
    if total_engines == 0:
        return "Unknown"
    
    malicious_percentage = (malicious_count / total_engines) * 100
    
    if malicious_percentage == 0:
        return "Safe"
    elif malicious_percentage <= 20:
        return "Low Risk"
    elif malicious_percentage <= 50:
        return "Moderate Risk"
    else:
        return "High Risk"

class IPCheckApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("IPCheck")
        self.setGeometry(100, 100, 700, 800)
        self.setStyleSheet("""
            QWidget { 
                background-color: black; 
                color: white; 
            }
            QLineEdit, QTextEdit {
                background-color: #333333; 
                color: white; 
                border: 1px solid red;
            }
            QPushButton {
                background-color: red; 
                color: white; 
                font-weight: bold;
                border: 1px solid white;
            }
            QTabBar::tab {
                background-color: red; 
                color: white;
                font-weight: bold;
                min-width: 150px;
                min-height: 40px;
                padding: 0px;
                margin: 2px;
                alignment: center;
            }
            QTabBar::tab:selected {
                background-color: darkred; 
            }
            QTabBar::tab:hover {
                background-color: #FF4500;
            }
            QProgressBar {
                background-color: #333333;
                color: white;
                border: 1px solid white;
                text-align: center;
            }
            QProgressBar::chunk {
                background-color: red;
            }
        """)

        self.selected_file_path = ""

        # Create tab widget
        self.notebook = QTabWidget(self)
        self.setCentralWidget(self.notebook)
        self.notebook.setTabBarAutoHide(False)
        self.notebook.setTabPosition(QTabWidget.North)

        # Create tabs
        self.main_tab = QWidget()
        self.api_tab = QWidget()
        self.hash_tab = QWidget()

        self.notebook.addTab(self.main_tab, "Reputation Check")
        self.notebook.addTab(self.api_tab, "API Key Management")
        self.notebook.addTab(self.hash_tab, "Hash Reputation")

        self.init_main_tab()
        self.init_api_tab()
        self.init_hash_tab()

        self.load_api_key()
    
    def init_main_tab(self):
        layout = QVBoxLayout(self.main_tab)
        
        # Single IP Lookup Section
        simple_ip_label = QLabel("Single IP Lookup")
        simple_ip_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(simple_ip_label)
        
        simple_ip_frame = QHBoxLayout()
        self.simple_ip_entry = QLineEdit()
        simple_ip_entry_button = QPushButton("Lookup")
        simple_ip_entry_button.clicked.connect(self.lookup_single_ip)
        
        simple_ip_frame.addWidget(self.simple_ip_entry)
        simple_ip_frame.addWidget(simple_ip_entry_button)
        layout.addLayout(simple_ip_frame)
        
        # Results display for single IP lookup
        self.single_lookup_result = QTextEdit()
        self.single_lookup_result.setReadOnly(True)
        layout.addWidget(self.single_lookup_result)
        
        # File selection
        self.file_label = QLabel("No file selected")
        layout.addWidget(self.file_label)
        
        file_button = QPushButton("Browse CSV/Excel File")
        file_button.clicked.connect(self.select_file)
        layout.addWidget(file_button)
        
        clear_button = QPushButton("Clear File")
        clear_button.clicked.connect(self.clear_file)
        layout.addWidget(clear_button)
        
        # Content text
        self.content_text = QTextEdit()
        layout.addWidget(self.content_text)
        
        # Progress bar
        self.progress_bar = QProgressBar()
        self.progress_bar.setAlignment(Qt.AlignCenter)  # Center the percentage text
        layout.addWidget(self.progress_bar)
        
        # Check button
        check_button = QPushButton("Check IP Reputation")
        check_button.clicked.connect(self.start_processing)
        layout.addWidget(check_button)
    
    def init_api_tab(self):
        outer_layout = QVBoxLayout(self.api_tab)
        layout = QVBoxLayout()
        layout.setAlignment(Qt.AlignCenter)
        
        api_key_label = QLabel("VirusTotal API Key:")
        api_key_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(api_key_label)
        
        self.api_key_entry = QLineEdit()
        self.api_key_entry.setEchoMode(QLineEdit.Password)
        layout.addWidget(self.api_key_entry)
        
        save_api_key_button = QPushButton("Save/Update API Key")
        save_api_key_button.clicked.connect(self.save_api_key)
        layout.addWidget(save_api_key_button)
        
        outer_layout.addLayout(layout)
    
    def init_hash_tab(self):
        layout = QVBoxLayout(self.hash_tab)
        
        # Create a tab widget inside the hash_tab
        self.hash_tab_widget = QTabWidget()
        layout.addWidget(self.hash_tab_widget)
        self.hash_tab_widget.setTabBarAutoHide(False)
        self.hash_tab_widget.setTabPosition(QTabWidget.North)

        # Apply the same stylesheet to the inner tab widget
        self.hash_tab_widget.setStyleSheet("""
            QTabBar::tab {
                background-color: red; 
                color: white;
                font-weight: bold;
                min-width: 150px;
                min-height: 40px;
                padding: 0px;
                margin: 2px;
                alignment: center;
            }
            QTabBar::tab:selected {
                background-color: darkred; 
            }
            QTabBar::tab:hover {
                background-color: #FF4500;
            }
        """)
        
        # Create the Hash Reputation tab
        hash_reputation_tab = QWidget()
        self.hash_tab_widget.addTab(hash_reputation_tab, "Reputation Check")
        hash_layout = QVBoxLayout(hash_reputation_tab)
        
        # Hash Reputation Section
        hash_label = QLabel("Hash Reputation Check")
        hash_label.setAlignment(Qt.AlignCenter)
        hash_layout.addWidget(hash_label)
        
        hash_frame = QHBoxLayout()
        self.hash_entry = QLineEdit()
        hash_lookup_button = QPushButton("Check Hash")
        hash_lookup_button.clicked.connect(self.lookup_hash)
        
        hash_frame.addWidget(self.hash_entry)
        hash_frame.addWidget(hash_lookup_button)
        hash_layout.addLayout(hash_frame)
        
        # Hash Results
        self.hash_lookup_result = QTextEdit()
        self.hash_lookup_result.setReadOnly(True)
        hash_layout.addWidget(self.hash_lookup_result)
        
        # Create the Details tab
        self.details_tab = QWidget()
        self.hash_tab_widget.addTab(self.details_tab, "Details")
        self.details_layout = QVBoxLayout(self.details_tab)
        self.details_text = QTextEdit()
        self.details_text.setReadOnly(True)
        self.details_layout.addWidget(self.details_text)
        
        # Add a button to search on the web
        self.search_button = QPushButton("Search on the Web")
        self.search_button.clicked.connect(self.search_on_web)
        self.details_layout.addWidget(self.search_button)
    
    def lookup_hash(self):
        api_key = self.api_key_entry.text().strip()
        file_hash = self.hash_entry.text().strip()
        
        if not api_key or not file_hash:
            QMessageBox.warning(self, "Warning", "Please enter API Key and File Hash.")
            return
        
        try:
            result = get_hash_reputation(api_key, file_hash)
            status = result[0]
            if status == "Unknown":
                QMessageBox.warning(self, "Warning", "Unable to retrieve hash reputation.")
                return
            elif status == "Not Found":
                QMessageBox.warning(self, "Warning", "File hash not found in VirusTotal database.")
                return
            elif status == "Invalid API Key":
                QMessageBox.warning(self, "Warning", "Invalid API Key.")
                return
            elif status == "Error":
                QMessageBox.warning(self, "Warning", "An error occurred while retrieving data.")
                return
            
            reputation, malicious_count, total_engines, community_score, file_signer, file_type, file_size, file_names = result
            
            # Determine community reputation
            if community_score == 0:
                community_reputation = "Safe"
            elif community_score == 1:
                community_reputation = "Suspicious"
            else:
                community_reputation = "Malicious"
            
            result_text = f"File Hash: {file_hash}\n"
            result_text += f"Reputation: {reputation}\n"
            result_text += f"Community Reputation: {community_reputation}\n"
            result_text += f"Malicious Detections: {malicious_count} out of {total_engines} engines"
            
            # Set text color based on reputation
            if reputation in ["High Risk", "Moderate Risk"] or community_reputation == "Malicious":
                self.hash_lookup_result.setStyleSheet("color: red;")
            else:
                self.hash_lookup_result.setStyleSheet("color: white;")
            
            self.hash_lookup_result.setText(result_text)
            
            # Update Details tab
            details_text = f"File Type: {file_type}\n"
            details_text += f"File Size: {file_size} bytes\n"
            details_text += f"Signer: {file_signer}\n"
            details_text += f"File Names: {', '.join(file_names)}\n"
            # Other details can be added here
            self.details_text.setText(details_text)
            
            # Store the file hash and names for the search button
            self.current_file_hash = file_hash
            self.current_file_names = file_names

        except Exception as e:
            QMessageBox.critical(self, "Error", f"An unexpected error occurred: {str(e)}")
    
    def search_on_web(self):
        if hasattr(self, 'current_file_names') and self.current_file_names:
            # Use the first file name for search
            query = self.current_file_names[0]
        else:
            # Use the file hash
            query = self.current_file_hash
        url = f"https://www.google.com/search?q={query}"
        webbrowser.open(url)
    
    def select_file(self):
        self.selected_file_path, _ = QFileDialog.getOpenFileName(self, "Select File", "", "CSV and Excel Files (*.csv *.xls *.xlsx)")
        if self.selected_file_path:
            self.file_label.setText(f"Selected File: {os.path.basename(self.selected_file_path)}")
            self.show_file_contents(self.selected_file_path)
        else:
            self.file_label.setText("No file selected")
    
    def clear_file(self):
        self.selected_file_path = ""
        self.file_label.setText("No file selected")
        self.content_text.clear()
    
    def show_file_contents(self, file_path):
        try:
            if file_path.endswith(".csv"):
                ip_df = pd.read_csv(file_path)
            elif file_path.endswith((".xls", ".xlsx")):
                ip_df = pd.read_excel(file_path)
            else:
                QMessageBox.critical(self, "Error", "Unsupported file format. Please select a CSV or Excel file.")
                return

            self.content_text.setText(ip_df.to_string(index=False))
        except Exception as e:
            self.content_text.setText(f"Error reading file: {e}")
    
    def lookup_single_ip(self):
        api_key = self.api_key_entry.text().strip()
        ip = self.simple_ip_entry.text().strip()
        
        if not api_key or not ip:
            QMessageBox.warning(self, "Warning", "Please enter API Key and IP address.")
            return

        try:
            isp, country, reputation = get_ip_info(api_key, ip)
            if isp is None:
                QMessageBox.critical(self, "Error", "Invalid API Key")
                return
            
            result_text = f"IP Address: {ip}\n"
            result_text += f"ISP: {isp}\n"
            result_text += f"Country: {country}\n"
            result_text += f"Reputation: {reputation}"
            
            # Set text color based on reputation
            if reputation == "Poor":
                self.single_lookup_result.setStyleSheet("color: red;")
            else:
                self.single_lookup_result.setStyleSheet("color: white;")
            
            self.single_lookup_result.setText(result_text)
        
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
    
    def start_processing(self):
        processing_thread = threading.Thread(target=self.process_csv)
        processing_thread.start()
    
    def process_csv(self):
        api_key = self.api_key_entry.text().strip()
        if not api_key:
            QMessageBox.warning(self, "Warning", "Please enter your VirusTotal API Key.")
            return

        if not self.selected_file_path:
            QMessageBox.warning(self, "Warning", "Please select a CSV or Excel file first.")
            return

        output_data = []
        try:
            if self.selected_file_path.endswith(".csv"):
                ip_df = pd.read_csv(self.selected_file_path)
            elif self.selected_file_path.endswith((".xls", ".xlsx")):
                ip_df = pd.read_excel(self.selected_file_path)
            else:
                QMessageBox.critical(self, "Error", "Unsupported file format. Please select a CSV or Excel file.")
                return

            if 'IP' not in ip_df.columns:
                QMessageBox.critical(self, "Error", "File must contain a column named 'IP'.")
                return
            
            self.progress_bar.setMaximum(len(ip_df))
            
            for idx, ip in enumerate(ip_df['IP']):
                isp, country, reputation = get_ip_info(api_key, ip)
                if isp is None:
                    QMessageBox.critical(self, "Error", "Invalid API Key")
                    return
                
                output_data.append({"S. No.": idx + 1, "IP": ip, "ISP": isp, "Country": country, "Reputation": reputation})
                self.progress_bar.setValue(idx + 1)

            output_df = pd.DataFrame(output_data)

            output_file, _ = QFileDialog.getSaveFileName(self, "Save As", downloads_folder, "Excel files (*.xlsx)")

            if not output_file:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Source IP Details"
            
            for r_idx, row in enumerate(dataframe_to_rows(output_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                                         top=Side(style="thin"), bottom=Side(style="thin"))
                    cell.border = thin_border
                    if r_idx == 1:
                        cell.fill = PatternFill(start_color="16365C", end_color="16365C", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)

            for column in ws.columns:
                max_length = max(len(str(cell.value)) for cell in column if cell.value)
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            wb.save(output_file)
            webbrowser.open(output_file)
        
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))
        
        self.progress_bar.setValue(0)
    
    def save_api_key(self):
        api_key = self.api_key_entry.text()
        if not api_key:
            QMessageBox.warning(self, "Warning", "Please enter an API Key before saving.")
            return
        try:
            with open(api_key_file, "w") as f:
                f.write(api_key)
            QMessageBox.information(self, "Success", "API Key saved successfully.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save API Key: {str(e)}")
    
    def load_api_key(self):
        if os.path.exists(api_key_file):
            with open(api_key_file, "r") as f:
                api_key = f.read().strip()
                self.api_key_entry.setText(api_key)

# Run the Application
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = IPCheckApp()
    window.show()
    sys.exit(app.exec_())
